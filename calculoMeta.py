from openpyxl import load_workbook

# Caminho e nome do arquivo excel com as metas de pagamento
excel_path = "metaExcel.xlsx"
excel_sheet = "metas"

# Caminho e nome do arquivo excel com os pagamentos realizados
excel_path2 = "caminho/do/arquivo/excel2.xlsx"
excel_sheet2 = "nome_da_folha2"

# Carrega o arquivo excel com as metas de pagamento
workbook = load_workbook(filename=excel_path, read_only=True)
worksheet = workbook[excel_sheet]

# Lê o valor da meta de pagamento
meta_de_pagamento = worksheet["A1"].value

# Carrega o arquivo excel com os pagamentos realizados
workbook2 = load_workbook(filename=excel_path2, read_only=True)
worksheet2 = workbook2[excel_sheet2]

# Lê os valores dos pagamentos realizados
valores_dos_pagamentos = [cell.value for cell in worksheet2["A"][1:]]

# Calcula a soma dos pagamentos realizados
soma_dos_pagamentos = sum(valores_dos_pagamentos)

# Calcula a porcentagem da meta que foi alcançada
porcentagem_alcancada = (soma_dos_pagamentos / meta_de_pagamento) * 100

# Exibe o resultado
print(f"A porcentagem da meta alcançada é de {porcentagem_alcancada}%.")
