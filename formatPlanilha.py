from openpyxl import load_workbook

# Carregar o arquivo Excel
workbook = load_workbook('terceirizadost.xlsx')

# Selecionar a planilha desejada
planilha = workbook['terceirizadost']

# Iterar pelas células da planilha
for row in planilha.iter_rows():
    for cell in row:
        # Verificar a formatação da célula
        if cell.font:
            print(f"Célula {cell.coordinate} - Cor da fonte: {cell.font.color.rgb}")
        if cell.fill:
            print(f"Célula {cell.coordinate} - Cor de fundo: {cell.fill.fgColor.rgb}")
        if cell.number_format:
            print(f"Célula {cell.coordinate} - Formato numérico: {cell.number_format}")

# Fechar o arquivo Excel
workbook.close()
